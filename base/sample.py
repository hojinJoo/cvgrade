from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
import json
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

TA_ID = '2022311491'
TA_PW = 'davidluiz4!'

SCORE_PAGE = "https://www.learnus.org/user/users.php?contextid=4311367&roleid=0&id=208457&perpage=5000&accesssince=0&search=&group=0&accesssince=0&roleid=5&search="


chrome_options = webdriver.ChromeOptions()
DRIVER = webdriver.Chrome(service=Service(
    ChromeDriverManager().install()), options=chrome_options)


def main():
    login()
    enter_student_info()
#students_info = get_students_info()
# toLearnUs(students_info)


def get_students_info():
    # TODO: 여기에 학생들 정보를 받아오는 코드를 짜면 됩니다.
    # data type은 dictionary 입니다.
    # key: student id
    # value: total_score and feedback content
    #
    # Example)
    # student_info = {
    #     '2021111111': {'total_score': '60.00', 'feedback': "너무 잘했어요"},
    #     '2021111112': {'total_score': '80.01', 'feedback': "아주 잘했어요"}
    # }

    students_info = {}

    wb = openpyxl.load_workbook(filename=xl)
    ws = wb.active
    for i in range(2, 18):
        st_id = ws.cell(row=i, column=1).value
        students_info[st_id] = {}
        students_info[st_id]['total_score'] = ws.cell(row=i, column=5).value
        students_info[st_id]['feedback'] = ws.cell(row=i, column=6).value

    # print(students_info)
    return students_info


def login():
    DRIVER.get('https://www.learnus.org/login.php')
    time.sleep(3)
    id_element = DRIVER.find_element(By.NAME, 'username')
    pw_element = DRIVER.find_element(By.NAME, 'password')

    id_element.send_keys(TA_ID)
    pw_element.send_keys(TA_PW)
    DRIVER.find_element(By.NAME, 'loginbutton').send_keys(Keys.ENTER)


def enter_student_info():
    l = []

    DRIVER.get(SCORE_PAGE)
    time.sleep(3)
    DRIVER.implicitly_wait(5)
    # print(By.CSS_SELECTOR())
    table = DRIVER.find_elements(
        By.CSS_SELECTOR, ".cell.c4.idnumber.txt-center.column-idnumber")
    # print(table)
    # print(table)
    for t in table:

        student_id = t.text
        # print(student_id)
        if student_id == "":
            break
        print(student_id)
        l.append(student_id)
        #t = d[student_id]
        #t['task_1_1'] = -1
        #t['task_1_2'] = -1
    #     t['task_2'] = -1

    # for student in table:
    #     try:
    #         time.sleep(1)
    #         td_elements = student.find_elements_by_tag_name('td')
    #         for td_element in td_elements:
    #             attribute_id = td_element.get_attribute("id")
    #             if "c3" in attribute_id:  # c3 is student_id

        # except Exception as E:
        #     print("Error raise: ", E)
    with open('student_ID_list.json', 'w') as w:
        json.dump(sorted(l), w, indent=4)


if __name__ == "__main__":
    main()
