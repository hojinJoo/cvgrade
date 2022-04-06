from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl


"""
이용법
SCORE_PAGE 변수에 채점 page를 넣으세요.
get_students_info() 함수만 수정하시면 됩니다.

미리 자기 아이디로 로그인 해서 페이지 맨 마지막에 있는
페이지당 과제들 : 모두
빠른 채점: 체크
로 해주세요.
"""

"""
함께 있는 크롬 드라이버는 맥용입니다. 윈도우에서 사용하실거면 바꾸세요.
런어스가 리다이렉팅이 좀 있어서 로드가 늦으면 잘 안될 때도 있습니다. 
100% 작동하게 코드 바꿔주시고 push 해주시는 분이 있으면 좋아요.
"""

TA_ID = '2022311491'
TA_PW = 'davidluiz4!'
CHROME_DRIVER_PATH = "./chromedriver.exe"
SCORE_PAGE = "https://www.learnus.org/mod/assign/view.php?id=2609275&action=grading"


DRIVER = webdriver.Chrome(CHROME_DRIVER_PATH)


def main():
    students_info = get_students_info()
    toLearnUs(students_info)


def get_students_info():
    # TODO: 여기에 학생들 정보를 받아오는 코드를 짜면 됩니다.
    # data type은 dictionary 입니다.
    # key: student id
    # value: total_score and feedback content
    #
    # Example)
    # student_info = {
    #     '2021111111': {'total_score': '60.00', 'feedback': "너무 잘했어요"},
    #     '2021111112': {'total_score': '80.01', 'feedback': "아주 잘했어요"}
    # }

    students_info = {}

    wb = openpyxl.load_workbook(filename=xl)
    ws = wb.active
    for i in range(2, 18):
        st_id = ws.cell(row=i, column=1).value
        students_info[st_id] = {}
        students_info[st_id]['total_score'] = ws.cell(row=i, column=5).value
        students_info[st_id]['feedback'] = ws.cell(row=i, column=6).value

    # print(students_info)
    return students_info


def toLearnUs(students_info):
    login()
    enter_student_info(students_info)


def login():
    DRIVER.get('https://www.learnus.org/login.php')
    time.sleep(3)
    id_element = DRIVER.find_element_by_name('username')
    pw_element = DRIVER.find_element_by_name('password')

    id_element.send_keys(TA_ID)
    pw_element.send_keys(TA_PW)
    DRIVER.find_element_by_name('loginbutton').send_keys(Keys.ENTER)


def enter_student_info(students_info):
    DRIVER.get(SCORE_PAGE)
    time.sleep(3)
    DRIVER.implicitly_wait(5)
    table = DRIVER.find_elements_by_tag_name('tr')
    time.sleep(3)
    for student in table:
        try:
            time.sleep(1)
            td_elements = student.find_elements_by_tag_name('td')
            for td_element in td_elements:
                attribute_id = td_element.get_attribute("id")
                if "c3" in attribute_id:  # c3 is student_id
                    student_id = td_element.text
                    # print(student_id)
                    # print(students_info[student_id])
                    continue

                if 'c6' in attribute_id and student_id in students_info.keys():  # c6 is score
                    score_area = td_element.find_element_by_tag_name('input')
                    time.sleep(1)
                    score_area.click()
                    score = students_info[student_id]['total_score']
                    # time.sleep(1)
                    score_area.send_keys(score)
                    # score_area.send_keys(Keys.ENTER)
                    continue

                if "c12" in attribute_id and student_id in students_info.keys():  # c12 is comment
                    text_area = td_element.find_element_by_tag_name('textarea')
                    time.sleep(1)
                    text_area.click()
                    comment = students_info[student_id]['feedback']
                    # time.sleep(1)
                    text_area.send_keys(comment)
                    # text_area.send_keys(Keys.ENTER)
                    continue

        except Exception as E:
            print("Error raise: ", E)


if __name__ == "__main__":
    main()
